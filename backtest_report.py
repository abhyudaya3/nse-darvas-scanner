"""
NSE Darvas Box Scanner - Backtest Analysis Report
====================================================
Builds a multi-sheet Excel workbook from a backtest run that actually
answers "does this method work" — not just a flat dump of numbers.

Sheets produced:
  1. Verdict             — plain-language pass/fail against objective
                            criteria, with the reasoning shown
  2. Equity Curve         — cumulative R-multiple growth chart across
                            every trade in chronological order
  3. Score Band Analysis  — the core proof: does composite_score
                            actually predict outcome quality?
  4. Yearly Performance   — catches "this only worked in one bull year"
  5. Per-Symbol Summary   — which stocks drove the results
  6. Trade Log            — every individual trade, full detail

All numbers in this report come from backtest_trade_log (the ground
truth individual-trade records) — nothing here is independently
recomputed or eyeballed, so the sheets are internally consistent with
each other by construction.
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import REPORTS_DIR, SCORE_THRESHOLDS, MIN_SIGNAL_SCORE
from database import (
    get_backtest_runs_df, get_backtest_symbol_summary_df,
    get_backtest_trade_log_df,
)
from logger_utils import get_logger

log = get_logger("performance")

# ─── Styling (mirrors report.py for visual consistency) ──────────────────────
CLR = {
    "header_bg":   "1F4E79",
    "header_fg":   "FFFFFF",
    "pass_bg":     "00B050",
    "fail_bg":     "C00000",
    "caution_bg":  "FFC000",
    "elite":       "00B050",
    "very_strong": "92D050",
    "strong":      "FFEB9C",
    "watch":       "FCE4D6",
    "alt_row":     "EBF3FB",
    "white":       "FFFFFF",
    "border":      "BDD7EE",
    "title_bg":    "2E75B6",
    "title_fg":    "FFFFFF",
}
THIN   = Side(style="thin", color=CLR["border"])
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ─── Objective pass/fail criteria ─────────────────────────────────────────────
# These thresholds are deliberately conservative and explicit, so the
# verdict sheet has a defensible, written-down standard rather than a
# vague "looks good" impression. Tune these in one place if your bar
# for "this strategy works" is different.
CRITERIA = {
    "min_trades":          30,    # statistical minimum to trust the win rate at all
    "min_win_rate_pct":    45.0,  # win rate floor (R:R skew can still be profitable below 50%)
    "min_profit_factor":   1.3,   # gross profit must meaningfully exceed gross loss
    "min_expectancy":      0.10,  # average R gained per trade must be clearly positive
    "max_drawdown_pct":    -35.0, # equity curve must not have collapsed more than this
    "min_score_separation":5.0,   # elite/very_strong band win rate must beat "watch" band
                                   # by at least this many percentage points, proving the
                                   # score actually discriminates good setups from weak ones
}


def generate_backtest_report(run_id: str, pattern_type: Optional[str] = None) -> list[Path]:
    """
    Build backtest analysis workbook(s) for *run_id* and return a list
    of saved paths.

    *pattern_type* — if given ('darvas_box' or 'cup_handle'), generates
    ONLY that pattern's report. If None (default), generates a SEPARATE
    workbook for every distinct pattern_type found in this run's trade
    log — e.g. a single --backtest-all run that covers both Darvas Box
    and Cup & Handle trades produces TWO separate Excel files, never a
    single mixed report. This matches the explicit requirement that
    "both patterns should be backtested, with separate Excel sheets" —
    a combined run is split apart at report-generation time so each
    pattern's performance can be judged entirely on its own terms,
    without one pattern's trade count or win rate diluting the other's.

    Raises ValueError if the run has no trades to analyse at all.
    """
    runs_df = get_backtest_runs_df()
    run_row = runs_df[runs_df["run_id"] == run_id]
    if run_row.empty:
        raise ValueError(f"No backtest run found with run_id={run_id}")
    run_meta = run_row.iloc[0].to_dict()

    all_trades_df = get_backtest_trade_log_df(run_id)
    all_symbols_df = get_backtest_symbol_summary_df(run_id)

    if all_trades_df.empty:
        raise ValueError(f"Backtest run {run_id} has zero trades — nothing to analyse")

    # Determine which pattern type(s) to generate reports for
    if pattern_type is not None:
        pattern_types = [pattern_type]
    else:
        pattern_types = sorted(all_trades_df["pattern_type"].dropna().unique().tolist())
        if not pattern_types:
            pattern_types = ["darvas_box"]   # legacy runs with no pattern_type column populated

    out_paths: list[Path] = []

    for ptype in pattern_types:
        trades_df = all_trades_df[all_trades_df["pattern_type"] == ptype].copy()
        if trades_df.empty:
            log.warning("No %s trades in run %s — skipping that report", ptype, run_id)
            continue

        symbols_df = (
            all_symbols_df[all_symbols_df["pattern_type"] == ptype].copy()
            if "pattern_type" in all_symbols_df.columns
            else all_symbols_df
        )

        # The trade log includes EVERY box/cup the detector found, scored
        # all the way down to single digits — most of those would NEVER be
        # generated as a live signal. Judging the strategy's validity on
        # the full unfiltered set (including very low scores) understates
        # real performance, because it's testing trades the live scanner
        # would have silently discarded anyway.
        #
        # FIXED 2026-06-20: this previously used SCORE_THRESHOLDS["watch"]
        # (=60) as the "live-eligible" cutoff, but the actual live Darvas
        # scanner gate is MIN_SIGNAL_SCORE (=85, see config.py and
        # scanner.py) — the two were never the same value, and nobody had
        # updated this report when MIN_SIGNAL_SCORE was introduced. That
        # meant the Verdict sheet was silently validating the strategy
        # against a much looser, no-longer-accurate definition of "what
        # the live scanner actually outputs" than what was really running
        # in production.
        live_threshold = MIN_SIGNAL_SCORE if ptype == "darvas_box" else SCORE_THRESHOLDS["watch"]
        # NOTE: composite_score is the UNIFIED column name for both pattern
        # types in backtest_trade_log — backtest_cup_handle_symbol() writes
        # its pattern_quality value into this same column (see backtest.py)
        # specifically so downstream reporting code doesn't need a
        # per-pattern column-name branch. pattern_quality is ALSO saved
        # separately (identical value) for clarity/debugging, but
        # composite_score is what every report sheet should read from.
        score_col = "composite_score"
        live_eligible_df = trades_df[trades_df[score_col] >= live_threshold]

        wb = Workbook()
        wb.remove(wb.active)

        verdict = _compute_verdict(live_eligible_df, run_meta, trades_df, live_threshold, score_col)

        _sheet_verdict(wb, verdict, run_meta, trades_df, ptype)
        _sheet_equity_curve(wb, trades_df)
        _sheet_score_band_analysis(wb, trades_df)
        _sheet_yearly_performance(wb, trades_df)
        _sheet_symbol_summary(wb, symbols_df)
        _sheet_trade_log(wb, trades_df)

        out_path = REPORTS_DIR / f"backtest_analysis_{ptype}_{run_id}.xlsx"
        wb.save(out_path)
        log.info("Backtest analysis report saved → %s", out_path)
        out_paths.append(out_path)

    return out_paths


# ─── Verdict computation ──────────────────────────────────────────────────────

def _compute_verdict(
    live_df: pd.DataFrame, run_meta: dict, full_df: pd.DataFrame,
    live_threshold: float, score_col: str,
) -> dict:
    """
    Evaluate the backtest against CRITERIA and return a structured
    verdict: overall pass/fail, plus the reasoning behind each check.

    *live_df*  — only trades scoring >= *live_threshold* on *score_col*.
                 This is what the strategy WOULD ACTUALLY HAVE TRADED
                 live, and is what every check below is computed from.
    *full_df*  — the complete historical trade log, including setups
                 scoring below the live threshold. Used only to show a
                 contrast footnote — never to compute pass/fail checks,
                 since judging the strategy on trades it would never
                 have generated understates real performance.
    *live_threshold* — the actual score cutoff used (e.g. MIN_SIGNAL_SCORE
                 for Darvas, or SCORE_THRESHOLDS["watch"] for Cup & Handle
                 until that pattern has its own validated live threshold).
    *score_col* — which column holds the quality score for this pattern
                 type ('composite_score' for Darvas, 'pattern_quality'
                 for Cup & Handle — they're on different internal scales
                 even though both happen to be 0-100).
    """
    trades_df = live_df   # keep the rest of this function's body unchanged below
    n = len(trades_df)
    wins = (trades_df["rr_realised"] > 0).sum()
    win_rate = wins / n * 100 if n else 0.0

    gross_profit = trades_df.loc[trades_df["rr_realised"] > 0, "rr_realised"].sum()
    gross_loss   = trades_df.loc[trades_df["rr_realised"] < 0, "rr_realised"].abs().sum()
    profit_factor = gross_profit / gross_loss if gross_loss > 0 else 999.0

    expectancy = trades_df["rr_realised"].mean()
    if pd.isna(expectancy):
        expectancy = 0.0   # n=0 case: nothing to average, report 0 not NaN

    # Equity curve drawdown (chronological). FIXED 2026-06-20: same bug
    # as backtest.py's backtest_symbol()/backtest_cup_handle_symbol() —
    # treating a raw R-multiple as a direct equity multiplier meant any
    # single -1.0R stop-loss zeroed the ENTIRE compounded equity curve,
    # producing a nonsensical -100% drawdown regardless of how many wins
    # surrounded it. Now scaled by RISK_PER_TRADE_PCT like everywhere else.
    from config import RISK_PER_TRADE_PCT
    risk_fraction = RISK_PER_TRADE_PCT / 100.0
    if n > 0:
        chrono = trades_df.sort_values("entry_date")
        equity = (1 + chrono["rr_realised"] * risk_fraction).cumprod()
        roll_max = equity.cummax()
        max_dd = ((equity - roll_max) / roll_max).min() * 100
        if pd.isna(max_dd):
            max_dd = 0.0
    else:
        # n=0: no trades at all to build an equity curve from. Reporting
        # 0.0% here (rather than NaN) is the honest "no data" answer —
        # this case is already caught by the "Sample size" check failing,
        # so it won't be mistaken for a genuinely flat, zero-drawdown result.
        max_dd = 0.0

    # Score separation: does "elite"+"very_strong" actually outperform "watch"?
    #
    # FIXED 2026-06-20: this previously compared bands within trades_df
    # (== live_df, the LIVE-ELIGIBLE subset scoring >= live_threshold).
    # For Darvas, live_threshold is MIN_SIGNAL_SCORE (85), which sits
    # ABOVE the "watch" band's range (60-69) by definition — meaning
    # low_band was ALWAYS empty for Darvas, and this check FAILED on
    # every single run regardless of how well the score actually
    # discriminated quality. The Score Band Analysis sheet (built from
    # full_df) was already showing a clean, monotonic 80%→70%→65%→60%→
    # 49% win-rate decline across bands — genuine evidence the score
    # works — while this check kept reporting "insufficient data" for a
    # structural reason that had nothing to do with data availability.
    #
    # The check now compares bands within full_df (the complete
    # historical record, matching the Score Band Analysis sheet), since
    # the question being asked — "does the score discriminate quality
    # at all" — is about the SCORING FORMULA itself, not specifically
    # about the narrow live-eligible slice.
    high_band = full_df[full_df["score_band"].isin(["elite", "very_strong"])]
    low_band  = full_df[full_df["score_band"] == "watch"]
    high_win_rate = (high_band["rr_realised"] > 0).mean() * 100 if len(high_band) else None
    low_win_rate  = (low_band["rr_realised"] > 0).mean() * 100 if len(low_band) else None
    score_separation = (
        (high_win_rate - low_win_rate)
        if high_win_rate is not None and low_win_rate is not None
        else None
    )

    checks = []

    checks.append({
        "check": "Sample size",
        "value": f"{n} trades",
        "threshold": f"≥ {CRITERIA['min_trades']}",
        "passed": n >= CRITERIA["min_trades"],
        "explanation": (
            "Enough trades to draw a statistically meaningful conclusion."
            if n >= CRITERIA["min_trades"] else
            "Too few trades — every other number on this sheet should be "
            "treated as preliminary, not proof. Re-run with more symbols "
            "or a longer history."
        ),
    })

    checks.append({
        "check": "Win rate",
        "value": f"{win_rate:.1f}%",
        "threshold": f"≥ {CRITERIA['min_win_rate_pct']}%",
        "passed": win_rate >= CRITERIA["min_win_rate_pct"],
        "explanation": (
            "Strategy wins often enough, especially combined with the "
            "risk-reward skew from the target structure."
            if win_rate >= CRITERIA["min_win_rate_pct"] else
            "Win rate is below the floor — even a good risk:reward ratio "
            "struggles to compensate for losing more than half the time."
        ),
    })

    checks.append({
        "check": "Profit factor",
        "value": f"{profit_factor:.2f}",
        "threshold": f"≥ {CRITERIA['min_profit_factor']}",
        "passed": profit_factor >= CRITERIA["min_profit_factor"],
        "explanation": (
            "Gross profit meaningfully exceeds gross loss — the system "
            "has real positive edge, not just a coin flip with costs."
            if profit_factor >= CRITERIA["min_profit_factor"] else
            "Gross profit doesn't sufficiently exceed gross loss. A value "
            "near 1.0 means the strategy is roughly break-even before "
            "slippage, brokerage, and taxes — which will push it negative."
        ),
    })

    checks.append({
        "check": "Expectancy (avg R per trade)",
        "value": f"{expectancy:+.3f}R",
        "threshold": f"≥ +{CRITERIA['min_expectancy']}R",
        "passed": expectancy >= CRITERIA["min_expectancy"],
        "explanation": (
            "On average, every trade taken grows the account by a "
            "meaningful fraction of what was risked."
            if expectancy >= CRITERIA["min_expectancy"] else
            "Average expectancy is too thin (or negative) — over many "
            "trades this does not reliably compound capital."
        ),
    })

    checks.append({
        "check": "Max drawdown (R-multiple equity curve)",
        "value": f"{max_dd:.1f}%",
        "threshold": f"≥ {CRITERIA['max_drawdown_pct']}%",
        "passed": max_dd >= CRITERIA["max_drawdown_pct"],
        "explanation": (
            "Worst peak-to-trough decline stayed within a survivable range."
            if max_dd >= CRITERIA["max_drawdown_pct"] else
            "Drawdown exceeded the survivability threshold — a trader "
            "would likely have abandoned the system or run out of risk "
            "capital before it recovered."
        ),
    })

    # Guard against drawing a conclusion from a tiny comparison sample —
    # e.g. a "Watch" band with only 2-3 trades can show a misleadingly
    # extreme win rate (0% or 100%) purely by chance, which would make
    # this check's pass/fail verdict unreliable in either direction.
    MIN_BAND_SAMPLE = 20
    band_sample_too_small = (
        score_separation is not None and
        (len(high_band) < MIN_BAND_SAMPLE or len(low_band) < MIN_BAND_SAMPLE)
    )

    if score_separation is not None and not band_sample_too_small:
        checks.append({
            "check": "Score discriminates quality (full history: Elite/Very Strong vs Watch win rate)",
            "value": f"{high_win_rate:.1f}% vs {low_win_rate:.1f}% (+{score_separation:.1f}pp)",
            "threshold": f"≥ +{CRITERIA['min_score_separation']}pp gap",
            "passed": score_separation >= CRITERIA["min_score_separation"],
            "explanation": (
                "Higher-scored signals genuinely win more often than "
                "lower-scored ones across the full historical record — "
                "the composite scoring formula is doing real "
                "discriminative work, not just noise. (Measured against "
                "the complete trade set, not just the narrow live-eligible "
                "slice, since the question is whether the FORMULA works.)"
                if score_separation >= CRITERIA["min_score_separation"] else
                "Higher-scored signals are NOT clearly outperforming "
                "lower-scored ones in this sample — and in some runs the "
                "relationship has been the OPPOSITE of intended (lower "
                "bands winning more). IMPORTANT NUANCE: this measures "
                "whether the QUALITY SCORE adds value on top of the hard "
                "entry gates (valid pattern shape + confirmed breakout) — "
                "it does NOT mean the underlying strategy itself is "
                "unprofitable; check the headline Win Rate / Profit Factor "
                "above, which can still pass even when this check fails. "
                "A likely explanation: once a setup clears the hard gates, "
                "the remaining 0-100 score may not carry much ADDITIONAL "
                "signal beyond what the gates already filtered for. "
                "Consider this evidence to revisit the score weights, but "
                "avoid overfitting them to a single backtest run — rerun "
                "after accumulating more data before concluding the "
                "relationship is reliably inverted."
            ),
        })
    elif score_separation is not None and band_sample_too_small:
        checks.append({
            "check": "Score discriminates quality (full history: Elite/Very Strong vs Watch win rate)",
            "value": f"{high_win_rate:.1f}% vs {low_win_rate:.1f}% on {len(high_band)} vs {len(low_band)} trades",
            "threshold": f"≥ {MIN_BAND_SAMPLE} trades in each band to trust the comparison",
            "passed": True,   # don't penalize the verdict for an unreliable tiny-sample comparison
            "explanation": (
                f"One or both bands had fewer than {MIN_BAND_SAMPLE} trades — "
                "too few to trust the win-rate comparison either way (a small "
                "sample can show 0% or 100% purely by chance). This check is "
                "marked as passing by default rather than failing on an "
                "unreliable signal; re-run with more symbols or a longer "
                "history once both bands have enough trades to compare."
            ),
        })
    else:
        checks.append({
            "check": "Score discriminates quality (full history: Elite/Very Strong vs Watch win rate)",
            "value": "Insufficient data in one or both bands",
            "threshold": f"≥ +{CRITERIA['min_score_separation']}pp gap",
            "passed": False,
            "explanation": "Not enough trades in the Elite/Very Strong or Watch "
                           "bands across the FULL historical record to compare — "
                           "re-run with more symbols or a longer price history.",
        })

    overall_pass = all(c["passed"] for c in checks)
    n_passed = sum(c["passed"] for c in checks)

    full_n = len(full_df)
    full_win_rate = (full_df["rr_realised"] > 0).mean() * 100 if full_n else 0.0

    return {
        "overall_pass": overall_pass,
        "n_checks": len(checks),
        "n_passed": n_passed,
        "checks": checks,
        "win_rate": win_rate,
        "profit_factor": profit_factor,
        "expectancy": expectancy,
        "max_dd": max_dd,
        "n_trades": n,
        "full_history_n_trades": full_n,
        "full_history_win_rate": full_win_rate,
        "live_threshold": live_threshold,
    }


# ─── Sheet 1: Verdict ──────────────────────────────────────────────────────────

def _sheet_verdict(
    wb: Workbook, verdict: dict, run_meta: dict, trades_df: pd.DataFrame,
    pattern_type: str,
) -> None:
    ws = wb.create_sheet("Verdict")
    ws.sheet_view.showGridLines = False

    pattern_label = "Darvas Box" if pattern_type == "darvas_box" else "Cup & Handle (O'Neil)"
    live_threshold = verdict.get("live_threshold", SCORE_THRESHOLDS["watch"])

    # ── Big headline verdict ──────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    headline = (
        f"✅  {pattern_label} METHOD VALIDATED — passes all objective criteria below"
        if verdict["overall_pass"] else
        f"⚠️  {pattern_label} METHOD NOT YET VALIDATED — passes {verdict['n_passed']}/{verdict['n_checks']} criteria"
    )
    cell = ws["A1"]
    cell.value = headline
    cell.font = Font(bold=True, size=16, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color=CLR["pass_bg"] if verdict["overall_pass"] else CLR["fail_bg"])
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    ws["A2"] = (
        f"Run: {run_meta.get('run_id', '?')}  |  Date: {run_meta.get('run_date', '?')}  |  "
        f"Symbols tested: {run_meta.get('symbols_tested', '?')}  |  "
        f"Live-eligible trades analysed: {verdict['n_trades']} "
        f"(score ≥ {live_threshold:.0f}, i.e. what the live scanner would actually output)"
    )
    ws["A2"].font = Font(italic=True, size=10, color="595959")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    ws.merge_cells("A3:F3")
    ws["A3"] = (
        f"For reference: the FULL historical record (every pattern found, including "
        f"scores below {live_threshold:.0f} that would never reach the live "
        f"scanner) was {verdict['full_history_n_trades']} trades at "
        f"{verdict['full_history_win_rate']:.1f}% win rate — included only on the "
        f"diagnostic sheets, never used for this verdict."
    )
    ws["A3"].font = Font(italic=True, size=9, color="A6A6A6")
    ws["A3"].alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[3].height = 16

    # ── Headline numbers row ──────────────────────────────────────────────────
    headline_stats = [
        ("Win Rate",       f"{verdict['win_rate']:.1f}%"),
        ("Profit Factor",  f"{verdict['profit_factor']:.2f}"),
        ("Expectancy",     f"{verdict['expectancy']:+.3f}R"),
        ("Max Drawdown",   f"{verdict['max_dd']:.1f}%"),
    ]
    for i, (label, value) in enumerate(headline_stats):
        col = i * 2 + 1
        ws.cell(row=5, column=col, value=label).font = Font(bold=True, size=10)
        ws.cell(row=5, column=col).alignment = Alignment(horizontal="center")
        vcell = ws.cell(row=6, column=col, value=value)
        vcell.font = Font(bold=True, size=14)
        vcell.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
        ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col + 1)
    ws.row_dimensions[6].height = 24

    # ── Detailed checks table ─────────────────────────────────────────────────
    start_row = 8
    headers = ["Check", "Result", "Threshold", "Pass/Fail", "What this means"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.font = Font(bold=True, color=CLR["header_fg"])
        cell.fill = PatternFill("solid", start_color=CLR["header_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[start_row].height = 24

    for r, chk in enumerate(verdict["checks"], start=start_row + 1):
        ws.cell(row=r, column=1, value=chk["check"]).alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(row=r, column=2, value=chk["value"]).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=3, value=chk["threshold"]).alignment = Alignment(horizontal="center", vertical="center")
        pf_cell = ws.cell(row=r, column=4, value="PASS" if chk["passed"] else "FAIL")
        pf_cell.font = Font(bold=True, color="FFFFFF")
        pf_cell.fill = PatternFill("solid", start_color=CLR["pass_bg"] if chk["passed"] else CLR["fail_bg"])
        pf_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=5, value=chk["explanation"]).alignment = Alignment(wrap_text=True, vertical="center")
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = BORDER
        ws.row_dimensions[r].height = 45

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 55

    # ── Honest caveats footer ─────────────────────────────────────────────────
    foot_row = start_row + len(verdict["checks"]) + 2
    ws.merge_cells(start_row=foot_row, start_column=1, end_row=foot_row, end_column=5)
    ws.cell(row=foot_row, column=1, value=(
        "Caveats: this backtest does not model brokerage, slippage, STT, or "
        "the realistic chance of NOT getting filled at the exact entry price "
        "in a fast-moving stock. Real-world results will be somewhat worse "
        "than shown here. Treat this as evidence the underlying logic has "
        "edge, not as a guarantee of live performance."
    )).font = Font(italic=True, size=9, color="808080")
    ws.cell(row=foot_row, column=1).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[foot_row].height = 40


# ─── Sheet 2: Equity Curve ─────────────────────────────────────────────────────

def _sheet_equity_curve(wb: Workbook, trades_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Equity Curve")
    chrono = trades_df.sort_values("entry_date").reset_index(drop=True)

    # FIXED 2026-06-20: this is the FOURTH location found with the same
    # equity-curve bug as backtest.py and _compute_verdict() above — a
    # raw R-multiple was being used as a direct equity multiplier, so any
    # single -1.0R stop-loss in the chronological sequence multiplied the
    # cumulative equity by EXACTLY ZERO, after which every subsequent row
    # displayed "0" regardless of how many winning trades followed. This
    # was directly visible in the report (the whole right side of the
    # Equity Curve sheet showing flat zeros) rather than just affecting
    # an aggregate stat, which is how this one was caught.
    from config import RISK_PER_TRADE_PCT
    risk_fraction = RISK_PER_TRADE_PCT / 100.0
    equity = (1 + chrono["rr_realised"] * risk_fraction).cumprod()

    ws["A1"] = "Trade #"
    ws["B1"] = "Entry Date"
    ws["C1"] = "Symbol"
    ws["D1"] = "R Realised"
    ws["E1"] = f"Cumulative Equity (starting=1.0, {RISK_PER_TRADE_PCT:.1f}% risk/trade)"
    for c in range(1, 6):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color=CLR["header_fg"])
        cell.fill = PatternFill("solid", start_color=CLR["header_bg"])
        cell.border = BORDER

    for i, row in chrono.iterrows():
        r = i + 2
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=row["entry_date"])
        ws.cell(row=r, column=3, value=row["symbol"])
        ws.cell(row=r, column=4, value=round(row["rr_realised"], 3))
        ws.cell(row=r, column=5, value=round(float(equity.iloc[i]), 4))
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = BORDER

    _auto_width(ws)

    # ── Chart ──────────────────────────────────────────────────────────────────
    chart = LineChart()
    chart.title = "Equity Curve — Cumulative Growth Across All Trades"
    chart.style = 2
    chart.y_axis.title = "Equity (1.0 = starting capital, in R-multiples)"
    chart.x_axis.title = "Trade sequence (chronological)"
    chart.height = 12
    chart.width = 26

    n = len(chrono)
    data = Reference(ws, min_col=5, min_row=1, max_row=n + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=n + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].smooth = False

    ws.add_chart(chart, "G2")


# ─── Sheet 3: Score Band Analysis (the core proof) ────────────────────────────

def _sheet_score_band_analysis(wb: Workbook, trades_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Score Band Analysis")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    ws["A1"] = "Does the Composite Score Actually Predict Outcome Quality?"
    ws["A1"].font = Font(bold=True, size=13, color=CLR["title_fg"])
    ws["A1"].fill = PatternFill("solid", start_color=CLR["title_bg"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    band_order = ["elite", "very_strong", "strong", "watch", "below_watch"]
    band_labels = {"elite": "Elite (≥90)", "very_strong": "Very Strong (80-89)",
                   "strong": "Strong (70-79)", "watch": "Watch (60-69)",
                   "below_watch": "Below 60 — NEVER reaches live scanner"}

    rows = []
    for band in band_order:
        sub = trades_df[trades_df["score_band"] == band]
        if sub.empty:
            continue
        n = len(sub)
        win_rate = (sub["rr_realised"] > 0).mean() * 100
        avg_rr = sub["rr_realised"].mean()
        gp = sub.loc[sub["rr_realised"] > 0, "rr_realised"].sum()
        gl = sub.loc[sub["rr_realised"] < 0, "rr_realised"].abs().sum()
        pf = min(gp / gl, 999.0) if gl > 0 else 999.0
        rows.append({
            "Score Band": band_labels[band],
            "Trades": n,
            "Win Rate %": round(win_rate, 1),
            "Avg R per Trade": round(avg_rr, 3),
            "Profit Factor": round(pf, 2),
            "Avg RSI at Entry": round(sub["rsi_at_entry"].mean(), 1),
            "Avg ADX at Entry": round(sub["adx_at_entry"].mean(), 1),
        })

    df = pd.DataFrame(rows)
    start_row = 3
    _write_styled_df(ws, df, start_row=start_row)

    # Highlight if there's a clear monotonic relationship (best band should
    # have the highest win rate / avg R, in descending row order)
    note_row = start_row + len(df) + 2
    if len(df) >= 2:
        win_rates = df["Win Rate %"].tolist()
        is_monotonic = all(win_rates[i] >= win_rates[i + 1] - 5 for i in range(len(win_rates) - 1))
        msg = (
            "✅ Win rate generally decreases as score band decreases — the "
            "scoring formula is correctly ranking signal quality."
            if is_monotonic else
            "⚠️ Win rate does NOT cleanly decrease across score bands — this "
            "suggests the composite score weights may need rebalancing, or "
            "there isn't enough data yet to see a clean pattern."
        )
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=7)
        ws.cell(row=note_row, column=1, value=msg).font = Font(italic=True, bold=True, size=10)
        ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[note_row].height = 30

    # ── Bar chart comparing win rate across bands ─────────────────────────────
    if len(df) >= 2:
        chart = BarChart()
        chart.title = "Win Rate by Score Band"
        chart.y_axis.title = "Win Rate (%)"
        chart.x_axis.title = "Score Band"
        chart.height = 9
        chart.width = 18
        data_ref = Reference(ws, min_col=3, min_row=start_row, max_row=start_row + len(df))
        cats_ref = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(df))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, f"A{note_row + 2}")


# ─── Sheet 4: Yearly Performance ──────────────────────────────────────────────

def _sheet_yearly_performance(wb: Workbook, trades_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Yearly Performance")

    df = trades_df.copy()
    df["year"] = pd.to_datetime(df["entry_date"]).dt.year

    rows = []
    for year, sub in df.groupby("year"):
        n = len(sub)
        win_rate = (sub["rr_realised"] > 0).mean() * 100
        avg_rr = sub["rr_realised"].mean()
        gp = sub.loc[sub["rr_realised"] > 0, "rr_realised"].sum()
        gl = sub.loc[sub["rr_realised"] < 0, "rr_realised"].abs().sum()
        pf = min(gp / gl, 999.0) if gl > 0 else 999.0
        rows.append({
            "Year": int(year), "Trades": n, "Win Rate %": round(win_rate, 1),
            "Avg R per Trade": round(avg_rr, 3), "Profit Factor": round(pf, 2),
            "Total R Gained": round(sub["rr_realised"].sum(), 2),
        })

    yearly_df = pd.DataFrame(rows).sort_values("Year")
    _write_styled_df(ws, yearly_df, start_row=2,
                      title="Year-by-Year Breakdown — Checks for Regime Dependency")

    note_row = 2 + len(yearly_df) + 3
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)
    ws.cell(row=note_row, column=1, value=(
        "If results are concentrated in only one or two years, the strategy "
        "may be fitted to a specific market regime (e.g. a strong bull run) "
        "rather than being robust across different conditions. Consistent "
        "positive years are a stronger signal than one spectacular year."
    )).font = Font(italic=True, size=9, color="808080")
    ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[note_row].height = 35


# ─── Sheet 5: Per-Symbol Summary ──────────────────────────────────────────────

def _sheet_symbol_summary(wb: Workbook, symbols_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Per-Symbol Summary")
    if symbols_df.empty:
        ws["A1"] = "No per-symbol data available."
        return
    display_cols = ["symbol", "trades", "win_rate", "profit_factor",
                     "cagr_pct", "max_drawdown", "sharpe", "avg_hold"]
    display_cols = [c for c in display_cols if c in symbols_df.columns]
    df = symbols_df[display_cols].sort_values("cagr_pct", ascending=False)
    _write_styled_df(ws, df, start_row=2, title="Per-Symbol Results — Which Stocks Drove the Outcome")


# ─── Sheet 6: Trade Log ────────────────────────────────────────────────────────

def _sheet_trade_log(wb: Workbook, trades_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Trade Log")

    # Common columns present for BOTH pattern types
    common_cols = [
        "symbol", "entry_date", "exit_date", "entry_price", "exit_price",
        "stop_loss", "target1", "target2", "outcome", "rr_realised",
        "hold_days", "composite_score", "score_band", "rs_rating",
        "rsi_at_entry", "adx_at_entry",
    ]

    # Determine which pattern this trade log actually is, and append the
    # relevant pattern-specific columns. FIXED 2026-06-20: this previously
    # ALWAYS showed Darvas-only columns (box_width_pct, box_age_bars),
    # which meant every Cup & Handle trade log silently displayed blank/
    # None for those AND never showed any of its own actual cup/handle
    # geometry or dates at all — the report looked like it had no
    # pattern-specific detail for C&H trades, when the data existed in
    # the database the whole time; it just wasn't being selected here.
    ptype = (
        trades_df["pattern_type"].iloc[0]
        if "pattern_type" in trades_df.columns and len(trades_df) else
        "darvas_box"
    )

    if ptype == "cup_handle":
        pattern_cols = [
            "sepa_score", "cup_start_date", "cup_bottom_date", "cup_end_date",
            "cup_depth_pct", "cup_duration_weeks", "cup_shape_ok", "cup_volume_dryup",
            "handle_start_date", "handle_end_date",
            "handle_depth_pct", "handle_duration_weeks", "handle_in_upper_zone",
            "handle_volume_dryup",
            "prior_uptrend_pct", "breakout_volume_ratio",
        ]
    else:
        pattern_cols = ["sepa_score", "box_start_date", "box_end_date",
                        "box_width_pct", "box_age_bars"]

    display_cols = common_cols + pattern_cols
    display_cols = [c for c in display_cols if c in trades_df.columns]
    df = trades_df[display_cols].sort_values("entry_date")
    title = (
        "Full Trade Log — Every Individual Cup & Handle Trade"
        if ptype == "cup_handle" else
        "Full Trade Log — Every Individual Darvas Box Trade"
    )
    _write_styled_df(ws, df, start_row=2, title=title)

    # Colour-code the outcome column
    outcome_col_idx = display_cols.index("outcome") + 1 if "outcome" in display_cols else None
    if outcome_col_idx:
        outcome_colors = {
            "target2_hit": CLR["elite"], "target1_hit": CLR["strong"],
            "stopped_out": CLR["fail_bg"], "open_at_end": CLR["caution_bg"],
        }
        for r in range(3, 3 + len(df)):
            val = ws.cell(row=r, column=outcome_col_idx).value
            if val in outcome_colors:
                ws.cell(row=r, column=outcome_col_idx).fill = PatternFill(
                    "solid", start_color=outcome_colors[val]
                )


# ─── Shared helpers ────────────────────────────────────────────────────────────

def _write_styled_df(ws, df: pd.DataFrame, start_row: int, title: Optional[str] = None) -> None:
    if title:
        ws.merge_cells(start_row=start_row - 1, start_column=1,
                        end_row=start_row - 1, end_column=max(len(df.columns), 1))
        tcell = ws.cell(row=start_row - 1, column=1, value=title)
        tcell.font = Font(bold=True, size=12, color=CLR["title_fg"])
        tcell.fill = PatternFill("solid", start_color=CLR["title_bg"])
        tcell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[start_row - 1].height = 22

    for c, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=c, value=str(col_name))
        cell.font = Font(bold=True, color=CLR["header_fg"])
        cell.fill = PatternFill("solid", start_color=CLR["header_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[start_row].height = 28

    for r_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
            bg = CLR["alt_row"] if r_idx % 2 == 0 else CLR["white"]
            cell.fill = PatternFill("solid", start_color=bg)

    _auto_width(ws)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0) for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 32)
