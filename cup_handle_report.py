"""
NSE Darvas Box Scanner - Cup and Handle Excel Report
=====================================================
Generates a professional multi-sheet Excel workbook dedicated to
Cup and Handle signals, separate from the Darvas Box report.

Sheets:
  1. Today's C&H Signals  — sorted by pattern quality × RS Rating
  2. Pattern Quality Guide — O'Neil's criteria, what each check means
  3. Watchlist             — active C&H signals being tracked
"""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.chart import LineChart
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import REPORTS_DIR
from logger_utils import get_logger

log = get_logger("scanner")

# ─── Styling ──────────────────────────────────────────────────────────────────
CLR = {
    "header_bg":    "1F4E79",
    "header_fg":    "FFFFFF",
    "breaking_out": "00B050",   # green  — price clearing pivot on big volume
    "near_pivot":   "FFC000",   # amber  — within 3% of pivot, waiting
    "watching":     "EBF3FB",   # light blue — still forming
    "alt_row":      "F2F7FB",
    "white":        "FFFFFF",
    "border":       "BDD7EE",
    "title_bg":     "2E75B6",
    "title_fg":     "FFFFFF",
    "ideal_cup":    "C6EFCE",   # green — cup depth/duration in O'Neil ideal range
    "ok_cup":       "FFEB9C",   # amber — valid but not ideal
    "weak_cup":     "FFC7CE",   # red   — outside ideal range
}
THIN   = Side(style="thin", color=CLR["border"])
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ─── Column definitions ───────────────────────────────────────────────────────
SIGNAL_COLS = [
    # (Header label, attribute name on CupHandleSignal, width, format hint)
    ("Symbol",               "symbol",               14, "text"),
    ("Sector",               "sector",               16, "text"),
    ("Status",               "status",               16, "text"),
    ("Timeframe",            "timeframe",            12, "text"),
    ("Other TFs Found",      "nested_pattern_count", 14, "num"),
    ("Current Price ₹",      "current_price",        14, "num"),
    ("Pivot Price ₹",        "pivot_price",          14, "num"),
    ("Buy Zone Low ₹",       "buy_zone_low",         14, "num"),
    ("Buy Zone High ₹",      "buy_zone_high",        14, "num"),
    ("Stop Loss ₹",          "stop_loss",            13, "num"),
    ("Target 1 ₹",           "target1",              12, "num"),
    ("Target 2 ₹",           "target2",              12, "num"),
    ("Target 3 ₹",           "target3",              12, "num"),
    ("R:R Ratio",            "rr_ratio",             10, "num"),
    ("Position Size",        "position_size",        13, "num"),
    ("Capital Reqd ₹",       "capital_required",     15, "num"),
    ("Risk ₹",               "risk_amount",          11, "num"),
    ("Cup Start Date",       "cup_start_date",       14, "text"),
    ("Cup Bottom Date",      "cup_bottom_date",      15, "text"),
    ("Cup End Date",         "cup_end_date",         14, "text"),
    ("Cup High ₹",           "cup_high",             12, "num"),
    ("Cup Low ₹",            "cup_low",              11, "num"),
    ("Cup Depth %",          "cup_depth_pct",        12, "num"),
    ("Cup Duration (wks)",   "cup_duration_weeks",   17, "num"),
    ("Cup Shape ✓",          "cup_shape_ok",         12, "bool"),
    ("Cup Vol Dry-up ✓",     "cup_volume_dryup",     15, "bool"),
    ("Handle Start Date",    "handle_start_date",    16, "text"),
    ("Handle End Date",      "handle_end_date",      15, "text"),
    ("Handle Depth %",       "handle_depth_pct",     14, "num"),
    ("Handle Wks",           "handle_duration_weeks",12, "num"),
    ("Handle Upper Zone ✓",  "handle_in_upper_zone", 18, "bool"),
    ("Handle Vol Dry-up ✓",  "handle_volume_dryup",  17, "bool"),
    ("Prior Uptrend %",      "prior_uptrend_pct",    15, "num"),
    ("ATR",                  "atr",                  8,  "num"),
    ("RS Rating",            "rs_rating",            10, "num"),
    ("SEPA Score",           "sepa_score",           11, "num"),
    ("Weekly Trend",         "weekly_trend",         13, "text"),
    ("Monthly Trend",        "monthly_trend",        13, "text"),
    ("Breakout Vol Ratio",   "breakout_vol_ratio",   17, "num"),
    ("Pattern Quality",      "pattern_quality",      15, "num"),
    ("Remarks",              None,                   30, "text"),
]


def generate_cup_handle_report(
    signals: list,
    today: Optional[date] = None,
) -> Path:
    """
    Build the Cup and Handle Excel report. *signals* is a list of
    CupHandleSignal objects from today's scan run.
    Returns the saved report path.
    """
    today = today or date.today()
    out_path = REPORTS_DIR / f"cup_handle_report_{today.isoformat()}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    _sheet_signals(wb, signals, today)
    _sheet_guide(wb)

    wb.save(out_path)
    log.info("Cup and Handle report saved → %s", out_path)
    return out_path


# ─── Sheet 1: Today's Signals ────────────────────────────────────────────────

def _sheet_signals(wb: Workbook, signals: list, today: date) -> None:
    ws = wb.create_sheet("C&H Signals")
    ws.sheet_view.showGridLines = False

    # Title
    n_cols = len(SIGNAL_COLS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tc = ws.cell(row=1, column=1,
                 value=f"NSE Cup and Handle Scanner (O'Neil Method) | {today.strftime('%A, %d %B %Y')}")
    tc.font  = Font(bold=True, size=13, color=CLR["title_fg"])
    tc.fill  = PatternFill("solid", start_color=CLR["title_bg"])
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Sub-title
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    sc = ws.cell(row=2, column=1,
                 value=f"Total signals: {len(signals)}  |  "
                       f"Breaking Out: {sum(1 for s in signals if s.is_breaking_out)}  |  "
                       f"Near Pivot (≥97%): {sum(1 for s in signals if s.status == 'Near Pivot')}  |  "
                       f"Watching: {sum(1 for s in signals if s.status == 'Watching')}")
    sc.font = Font(italic=True, size=10, color="595959")
    sc.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # Header row
    for c_idx, (label, _, width, _) in enumerate(SIGNAL_COLS, start=1):
        cell = ws.cell(row=3, column=c_idx, value=label)
        cell.font = Font(bold=True, color=CLR["header_fg"], size=10)
        cell.fill = PatternFill("solid", start_color=CLR["header_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c_idx)].width = width
    ws.row_dimensions[3].height = 32

    # Sort: Breaking Out first, then Near Pivot, then Watching; within each
    # group sort by pattern_quality × rs_rating descending (composite rank)
    def sort_key(s):
        order = {"Breaking Out": 0, "Near Pivot": 1, "Watching": 2}
        return (order.get(s.status, 9), -(s.pattern_quality * s.rs_rating))

    sorted_sigs = sorted(signals, key=sort_key)

    for r_idx, sig in enumerate(sorted_sigs, start=4):
        bg = (CLR["alt_row"] if r_idx % 2 == 0 else CLR["white"])

        for c_idx, (_, attr, _, fmt) in enumerate(SIGNAL_COLS, start=1):
            if attr is None:
                val = _auto_remark_ch(sig)
            elif fmt == "bool":
                val = "✅" if getattr(sig, attr) else "❌"
            else:
                val = getattr(sig, attr, "")

            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Status-based row color override on the Status column (col 3)
            if c_idx == 3:
                if sig.status == "Breaking Out":
                    cell.fill = PatternFill("solid", start_color=CLR["breaking_out"])
                    cell.font = Font(bold=True, color="FFFFFF")
                elif sig.status == "Near Pivot":
                    cell.fill = PatternFill("solid", start_color=CLR["near_pivot"])
                    cell.font = Font(bold=True)
                else:
                    cell.fill = PatternFill("solid", start_color=CLR["watching"])
            else:
                cell.fill = PatternFill("solid", start_color=bg)

        # Color-code cup depth column (col 18) by O'Neil ideal ranges
        depth_col = 18
        depth_cell = ws.cell(row=r_idx, column=depth_col)
        d = sig.cup_depth_pct
        if 15 <= d <= 25:
            depth_cell.fill = PatternFill("solid", start_color=CLR["ideal_cup"])
        elif 12 <= d <= 33:
            depth_cell.fill = PatternFill("solid", start_color=CLR["ok_cup"])
        else:
            depth_cell.fill = PatternFill("solid", start_color=CLR["weak_cup"])

        ws.row_dimensions[r_idx].height = 18

    ws.freeze_panes = "A4"


# ─── Sheet 2: O'Neil Quality Guide ───────────────────────────────────────────

def _sheet_guide(wb: Workbook) -> None:
    ws = wb.create_sheet("O'Neil Quality Guide")
    ws.sheet_view.showGridLines = False

    rows = [
        ("CRITERIA", "O'NEIL'S RULE", "IDEAL RANGE", "VALID RANGE", "WHY IT MATTERS"),

        ("Prior Uptrend",
         "Stock must be a leader BEFORE the base. The cup forms as a "
         "rest after a significant advance.",
         "≥ 30% over 6 months",
         "≥ 30%",
         "A cup in a weak/falling stock is not a valid O'Neil base — it's "
         "just a sideways correction. The prior run confirms institutional "
         "sponsorship exists before the base."),

        ("Cup Depth",
         "How far price declines from the cup's left-side high to its bottom.",
         "15–25%",
         "12–33% (up to 50% in severe market corrections)",
         "Too shallow (<12%) and the base didn't shake out weak holders. "
         "Too deep (>33%) and supply-demand dynamics are broken. The ideal "
         "15-25% range represents a normal, healthy rest."),

        ("Cup Duration",
         "How many weeks the entire cup (left side + bottom + right side) spans.",
         "12–26 weeks (3–6 months)",
         "7–65 weeks",
         "Too short (< 7 weeks) and there wasn't enough time for a proper "
         "base to form. The most reliable cups run 3-6 months. Very long "
         "cups (>65 weeks) suggest a structural problem, not just a rest."),

        ("Cup Shape",
         "The cup should form a rounded 'U' bottom — gradual decline, "
         "rounding base, gradual recovery.",
         "Rounded U, declining volume into bottom",
         "No sharp V-bottoms",
         "A V-shaped bottom means panic selling followed by panic buying — "
         "institutions aren't building positions methodically. A rounded U "
         "shows controlled, orderly accumulation near the base."),

        ("Volume at Cup Bottom",
         "Volume should dry up as price approaches the cup bottom, then "
         "pick up as price climbs the right side.",
         "Volume 20-30%+ below average near bottom",
         "Below-average volume near bottom",
         "Volume dry-up near the low confirms selling is exhausted — no "
         "more supply pressure. Rising volume on the right side confirms "
         "demand is returning."),

        ("Handle Position",
         "The handle must form in the UPPER half of the cup — NOT in the "
         "lower half.",
         "Upper third of cup range",
         "Upper half of cup range",
         "A handle that drops into the lower half of the cup is called a "
         "'faulty handle' by O'Neil — it means the stock is too weak to "
         "hold its gains and likely has unresolved overhead supply."),

        ("Handle Depth",
         "The handle's own pullback, measured from its high to its low.",
         "5–8%",
         "Up to 12%",
         "Handles that drop more than 12% from their own high are shaking "
         "out too many strong holders and often fail at the pivot. Shallow "
         "handles (5-8%) on light volume are the most bullish sign."),

        ("Handle Volume",
         "Volume should decline during the handle formation, showing "
         "sellers are drying up.",
         "40-50% below cup average",
         "Below cup average",
         "Declining volume in the handle confirms supply is exhausted "
         "and the stock is coiling for a breakout. Heavy volume during "
         "the handle is a warning sign of institutional distribution."),

        ("Breakout Volume",
         "The day price clears the handle's high (the pivot point), "
         "volume must surge significantly.",
         "50%+ above 50-day average",
         "40%+ above 50-day average",
         "A low-volume breakout is a false start — it means institutions "
         "aren't buying. O'Neil insisted on 40-50% above-average volume "
         "on the breakout day to confirm genuine institutional demand."),

        ("Buy Zone",
         "The optimal buy point is the pivot (handle high) up to 5% "
         "above it — O'Neil's 'buy zone'.",
         "Exactly at pivot to +3% above",
         "Pivot to +5% above",
         "Chasing a stock more than 5% above the pivot means you're buying "
         "after institutions already have their full position — you'll be "
         "shaken out on the first normal pullback to the pivot."),

        ("RS Rating",
         "The stock's price performance relative to all other stocks "
         "(O'Neil's Relative Strength Rating, 1-99 percentile).",
         "RS ≥ 90 at time of breakout",
         "RS ≥ 80",
         "O'Neil's research showed most big market winners had an RS "
         "Rating of 80+ BEFORE they broke out. An RS below 70 at the "
         "pivot is a major warning sign — the stock is already lagging "
         "the market even before it's fully broken out."),
    ]

    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = "William O'Neil Cup and Handle — Quality Criteria Reference Guide"
    title_cell.font  = Font(bold=True, size=13, color=CLR["title_fg"])
    title_cell.fill  = PatternFill("solid", start_color=CLR["title_bg"])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    col_widths = [22, 40, 22, 30, 55]
    for c, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    for r_idx, row_data in enumerate(rows, start=2):
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if r_idx == 2:  # header
                cell.font = Font(bold=True, color=CLR["header_fg"])
                cell.fill = PatternFill("solid", start_color=CLR["header_bg"])
            elif r_idx % 2 == 0:
                cell.fill = PatternFill("solid", start_color=CLR["alt_row"])
            else:
                cell.fill = PatternFill("solid", start_color=CLR["white"])
        ws.row_dimensions[r_idx].height = 55

    ws.freeze_panes = "A3"


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _auto_remark_ch(sig) -> str:
    parts = []
    if sig.is_breaking_out:
        parts.append("🚀 BREAKOUT — buy now within buy zone")
    if sig.rs_rating >= 90:
        parts.append("RS Leader (≥90)")
    elif sig.rs_rating >= 80:
        parts.append("RS Strong (≥80)")
    if sig.cup_depth_pct <= 25 and sig.cup_depth_pct >= 15:
        parts.append("Ideal cup depth")
    if sig.cup_volume_dryup and sig.handle_volume_dryup:
        parts.append("Vol dry-up ✓✓")
    if sig.weekly_trend == "bullish" and sig.monthly_trend == "bullish":
        parts.append("MTF bullish")
    if sig.sepa_score >= 8:
        parts.append("SEPA compliant")
    if not parts:
        parts.append("Cup forming")
    return " | ".join(parts)
