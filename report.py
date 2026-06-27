"""
NSE Darvas Box Scanner - Excel Report Generator
=================================================
Produces a professional multi-sheet Excel workbook:
  Sheet 1 – Today's Signals   (sorted by composite score)
  Sheet 2 – Active Watchlist
  Sheet 3 – Strategy Performance
  Sheet 4 – Adaptive Analysis
"""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

from config import REPORTS_DIR, SCORE_THRESHOLDS
from database import get_all_signals_df, get_watchlist_df, get_performance_df
from logger_utils import get_logger
from signal_tracker import adaptive_analysis

log = get_logger("scanner")

# ─── Colour palette ───────────────────────────────────────────────────────────
CLR = {
    "header_bg":  "1F4E79",  # dark navy
    "header_fg":  "FFFFFF",
    "elite":      "00B050",  # green
    "very_strong":"92D050",
    "strong":     "FFEB9C",
    "weak":       "FFC7CE",
    "alt_row":    "EBF3FB",
    "white":      "FFFFFF",
    "border":     "BDD7EE",
    "title_bg":   "2E75B6",
    "title_fg":   "FFFFFF",
}

THIN = Side(style="thin", color=CLR["border"])
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def generate_report(signals: list, today: Optional[date] = None) -> Path:
    """
    Build the Excel report and return its path.
    *signals* is a list of scanner.Signal objects from today's run.
    """
    today = today or date.today()
    out_path = REPORTS_DIR / f"darvas_report_{today.isoformat()}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    _sheet_signals(wb, signals)
    _sheet_watchlist(wb)
    _sheet_performance(wb)
    _sheet_adaptive(wb)

    wb.save(out_path)
    log.info("Excel report saved → %s", out_path)
    return out_path


# ─── Sheet 1: Today's Signals ────────────────────────────────────────────────

SIGNAL_COLS = [
    ("Symbol",          "symbol"),
    ("Sector",          "sector"),
    ("Current Price",   "current_price"),
    ("Box High",        "box_high"),
    ("Box Low",         "box_low"),
    ("Box Start Date",  "box_start_date"),
    ("Box End Date",    "box_end_date"),
    ("Entry Zone Low",  "entry_zone_low"),
    ("Entry Zone High", "entry_zone_high"),
    ("Stop Loss",       "stop_loss"),
    ("Target 1",        "target1"),
    ("Target 2",        "target2"),
    ("Target 3",        "target3"),
    ("ATR",             "atr"),
    ("Position Size",   "position_size"),
    ("Capital (INR)",   "capital_required"),
    ("Risk/Share",      "risk_per_share"),
    ("R:R Ratio",       "rr_ratio"),
    ("RSI",             "rsi_val"),
    ("ADX",             "adx_val"),
    ("Vol Ratio",       "volume_ratio"),
    ("Weekly Trend",    "weekly_trend"),
    ("Monthly Trend",   "monthly_trend"),
    ("RS Rating",       "rs_rating"),
    ("SEPA Score",      "sepa_score"),
    ("Composite Score", "composite_score"),
    ("Classification",  "classification"),
    ("Remarks",         None),
]


def _sheet_signals(wb: Workbook, signals: list) -> None:
    ws = wb.create_sheet("Today's Signals")
    _title_row(ws, f"NSE Darvas Box Scanner – Bottom-of-Box Setups  |  {date.today()}", len(SIGNAL_COLS))

    # Header
    _header_row(ws, [c[0] for c in SIGNAL_COLS], row=2)

    # Sort signals by composite score
    sorted_sigs = sorted(signals, key=lambda s: s.composite_score, reverse=True)

    for r_idx, sig in enumerate(sorted_sigs, start=3):
        for c_idx, (_, attr) in enumerate(SIGNAL_COLS, start=1):
            val = getattr(sig, attr, "") if attr else _auto_remark(sig)
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Alternate row fill
            bg = CLR["alt_row"] if r_idx % 2 == 0 else CLR["white"]
            cell.fill = PatternFill("solid", start_color=bg)

        # Score-based row highlight in classification column
        score = sig.composite_score
        last_col = len(SIGNAL_COLS)
        cl_cell = ws.cell(row=r_idx, column=last_col - 1)   # Classification col
        if score >= SCORE_THRESHOLDS["elite"]:
            cl_cell.fill = PatternFill("solid", start_color=CLR["elite"])
            cl_cell.font = Font(bold=True, color="FFFFFF")
        elif score >= SCORE_THRESHOLDS["very_strong"]:
            cl_cell.fill = PatternFill("solid", start_color=CLR["very_strong"])
        elif score >= SCORE_THRESHOLDS["strong"]:
            cl_cell.fill = PatternFill("solid", start_color=CLR["strong"])

    _auto_width(ws)
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False


# ─── Sheet 2: Watchlist ──────────────────────────────────────────────────────

def _sheet_watchlist(wb: Workbook) -> None:
    ws = wb.create_sheet("Watchlist")
    df = get_watchlist_df()
    if df.empty:
        ws["A1"] = "No active watchlist entries."
        return
    _title_row(ws, "Active Watchlist", len(df.columns))
    _write_df(ws, df, start_row=2)


# ─── Sheet 3: Performance ─────────────────────────────────────────────────────

def _sheet_performance(wb: Workbook) -> None:
    ws = wb.create_sheet("Performance")
    df = get_performance_df()
    if df.empty:
        ws["A1"] = "No performance data yet – signals are still being tracked."
        return
    _title_row(ws, "Strategy Effectiveness by Score Band", len(df.columns))
    _write_df(ws, df, start_row=2)


# ─── Sheet 4: Adaptive Analysis ──────────────────────────────────────────────

def _sheet_adaptive(wb: Workbook) -> None:
    ws = wb.create_sheet("Adaptive Analysis")
    df = adaptive_analysis()
    if df.empty:
        ws["A1"] = "Insufficient data for adaptive analysis (need ≥20 triggered signals)."
        return
    _title_row(ws, "Factor-Bin Success Rates", len(df.columns))
    _write_df(ws, df, start_row=2)


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _title_row(ws, title: str, n_cols: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font  = Font(bold=True, size=13, color=CLR["title_fg"])
    cell.fill  = PatternFill("solid", start_color=CLR["title_bg"])
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22


def _header_row(ws, headers: list[str], row: int = 2) -> None:
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font  = Font(bold=True, color=CLR["header_fg"])
        cell.fill  = PatternFill("solid", start_color=CLR["header_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 30


def _write_df(ws, df: pd.DataFrame, start_row: int = 2) -> None:
    _header_row(ws, list(df.columns), row=start_row)
    for r_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
            bg = CLR["alt_row"] if r_idx % 2 == 0 else CLR["white"]
            cell.fill = PatternFill("solid", start_color=bg)
    _auto_width(ws)
    ws.freeze_panes = f"A{start_row + 1}"


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0) for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)


def _auto_remark(sig) -> str:
    parts = []
    if sig.rs_rating >= 90:
        parts.append("RS Leader")
    if sig.sepa_score >= 8:
        parts.append("SEPA Compliant")
    if sig.weekly_trend == "bullish" and sig.monthly_trend == "bullish":
        parts.append("MTF Bullish")
    if sig.volume_ratio >= 2.0:
        parts.append("Vol Surge")
    return " | ".join(parts) if parts else "Standard Setup"
